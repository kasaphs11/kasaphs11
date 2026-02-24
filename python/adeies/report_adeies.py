# -*- coding: utf-8 -*-
"""
Report Adeies (GUI)
-------------------
Ζητούμενα:
- Υποστήριξη πολλαπλών ΑΣΜ (με κενό / κόμμα / ; / new line / tab).
- Default επιλογή στο GUI: ΑΣΜ.
- Πεδίο Θητεία (9/8/6) σε ξεχωριστή γραμμή, και από κάτω η Αναζήτηση.
- Για ΑΣΜ mode: παράγει ένα Excel με blocks ανά άτομο.
  Δομή block:
    1) Γραμμή ονόματος (bold, ΟΧΙ συγχώνευση)
    2) Headers
    3) Δεδομένα
    4) Γραμμή "Σύνολο" με bold στο label και στα νούμερα (ίδια γραμμή)
    5) Γραμμή "Υπόλοιπο" κάτω ακριβώς από το "Σύνολο" (bold) με βάση τη Θητεία
    6) Περίγραμμα (outline) γύρω από headers+δεδομένα+σύνολο+υπόλοιπο για εύκολο διαχωρισμό.
- Οι στήλες που κρατάμε είναι ΜΟΝΟ οι παρακάτω (αν υπάρχουν στο αρχικό excel).
- Οι στήλες με 'ΗΜ/ΝΙΑ' εμφανίζονται ως DD/MM/YYYY (χωρίς ώρα).
"""

import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

APP_TITLE = "Report Adeies"

# ΜΟΝΟ αυτές τις στήλες κρατάμε (όπως στο παράδειγμα)
KEEP_COLUMNS = [
    "Α/Α",
    "ΠΑΡΑΤΗΡΗΣΕΙΣ",  # Αν θες να ΜΗΝ κρατιέται, βάλτο στο REMOVE_COLUMNS
    "ΗΜ/ΝΙΑ ΕΝΑΡΞΗΣ",
    "ΗΜ/ΝΙΑ ΛΗΞΗΣ",
    "ΚΑ",
    "ΤΑΠ",
    "ΕΛΔΥΚ",
    "ΤΙΜ",
    "ΑΙΜ",
    "ΑΓΡ",
    "ΦΟΙΤ",
    "ΑΝΑΡ",
    "ΗΜΕΡΕΣ",
]

# Αν θες να αφαιρούνται πάντα
REMOVE_COLUMNS = {"ΠΑΡΑΤΗΡΗΣΕΙΣ"}

# Ποιες αριθμητικές στήλες αθροίζουμε
NUM_SUM_COLUMNS = ["ΚΑ", "ΤΑΠ", "ΕΛΔΥΚ", "ΤΙΜ", "ΑΙΜ", "ΑΓΡ", "ΦΟΙΤ", "ΑΝΑΡ", "ΗΜΕΡΕΣ"]


def safe_str(x):
    return "" if x is None else str(x).strip()


def parse_multi_values(text):
    """
    Parse multiple values (ΑΣΜ) from a single input field.
    Accepts separators: spaces, commas, semicolons, new lines, tabs.
    Returns a list preserving order (duplicates removed).
    """
    if text is None:
        return []
    raw = str(text).strip()
    if not raw:
        return []
    parts = [p for p in re.split(r"[\s,;]+", raw) if p]

    # de-duplicate while preserving order
    seen = set()
    out = []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def detect_headers(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = safe_str(ws.cell(row=1, column=col).value)
        if val:
            headers[val.upper()] = col
    return headers


def to_date(value):
    """Try to convert to Python date/datetime; return value unchanged if not possible."""
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return value


def apply_outline_border(ws, min_row, max_row, min_col, max_col, outline_style="medium", inner_style="thin"):
    """Apply a border outline around a rectangular block, with optional inner borders."""
    if min_row > max_row or min_col > max_col:
        return

    side_outline = Side(style=outline_style)
    side_inner = Side(style=inner_style)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)

            left = side_outline if c == min_col else side_inner
            right = side_outline if c == max_col else side_inner
            top = side_outline if r == min_row else side_inner
            bottom = side_outline if r == max_row else side_inner

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("720x360")
        self.resizable(False, False)

        self.input_path = tk.StringVar()
        self.query = tk.StringVar()
        # ✅ Προεπιλογή ΑΣΜ
        self.search_mode = tk.StringVar(value="asm")
        # Θητεία (default 9)
        self.thiteia = tk.StringVar(value="9")

        self.build_ui()

    def build_ui(self):
        frm = ttk.Frame(self, padding=12)
        frm.pack(fill="both", expand=True)

        row = 0
        ttk.Label(frm, text="Αρχείο Excel:", font=("Segoe UI", 10, "bold")).grid(row=row, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.input_path, width=70).grid(row=row, column=1, padx=8)
        ttk.Button(frm, text="Επιλογή...", command=self.pick_file).grid(row=row, column=2)

        row += 1
        lf = ttk.LabelFrame(frm, text="Τύπος Αναζήτησης", padding=(8, 6))
        lf.grid(row=row, column=0, columnspan=3, sticky="we", pady=(10, 0))
        ttk.Radiobutton(
            lf, text="Ονοματεπώνυμο (ΕΠΩΝΥΜΟ + ΟΝΟΜΑ)", value="name", variable=self.search_mode
        ).grid(row=0, column=0, sticky="w")
        ttk.Radiobutton(lf, text="ΑΣΜ", value="asm", variable=self.search_mode).grid(
            row=0, column=1, sticky="w", padx=(10, 0)
        )

        # ✅ Ξεχωριστή γραμμή Θητεία
        row += 1
        ttk.Label(frm, text="Θητεία:", font=("Segoe UI", 10)).grid(row=row, column=0, sticky="w", pady=(10, 0))
        th = ttk.Combobox(frm, textvariable=self.thiteia, width=10, state="readonly", values=("9", "8", "6"))
        th.grid(row=row, column=1, sticky="w", padx=8, pady=(10, 0))
        th.current(0)

        # ✅ Από κάτω η Αναζήτηση όπως πριν
        row += 1
        ttk.Label(frm, text="Αναζήτηση:", font=("Segoe UI", 10)).grid(row=row, column=0, sticky="w", pady=(10, 0))
        ttk.Entry(frm, textvariable=self.query, width=60).grid(row=row, column=1, padx=8, pady=(10, 0))
        ttk.Button(frm, text="Δημιουργία Excel", command=self.run).grid(row=row, column=2, pady=(10, 0))

        row += 1
        help_text = (
            "• Για ΑΣΜ μπορείς να βάλεις πολλά: π.χ. 2050... 2530... ή 2050...,2530...\n"
            "• Θα κρατηθούν ΜΟΝΟ οι στήλες του report (όπως στη φωτο), με ημερομηνίες DD/MM/YYYY.\n"
            "• Θα σου ζητηθεί πού να αποθηκευτεί το νέο Excel."
        )
        ttk.Label(frm, text=help_text, foreground="#444").grid(
            row=row, column=0, columnspan=3, sticky="w", pady=(14, 0)
        )

    def pick_file(self):
        fpath = filedialog.askopenfilename(
            title="Διάλεξε Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm")],
        )
        if fpath:
            self.input_path.set(fpath)

    def run(self):
        try:
            path = self.input_path.get().strip()
            if not path:
                messagebox.showwarning(APP_TITLE, "Διάλεξε αρχείο Excel.")
                return

            q = self.query.get().strip()
            if not q:
                messagebox.showwarning(APP_TITLE, "Γράψε τι θα αναζητήσουμε (Ονοματεπώνυμο ή ΑΣΜ).")
                return

            out_path = self.process(path, q, self.search_mode.get(), self.thiteia.get())
            if out_path:
                messagebox.showinfo(APP_TITLE, f"Αποθηκεύτηκε:\n{out_path}")
        except Exception as e:
            import traceback

            traceback.print_exc()
            messagebox.showerror(APP_TITLE, f"Σφάλμα:\n{e}")

    def process(self, excel_path, query, mode, thiteia):
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        headers = detect_headers(ws)

        # Allowance βάση θητείας
        th_map = {
            "9": {"ΚΑ": 15, "ΤΑΠ": 10, "ΕΛΔΥΚ": 5},
            "8": {"ΚΑ": 12, "ΤΑΠ": 8, "ΕΛΔΥΚ": 5},
            "6": {"ΚΑ": 9, "ΤΑΠ": 6, "ΕΛΔΥΚ": 5},
        }
        base_allowance = th_map.get(str(thiteia).strip(), th_map["9"])

        col_eponymo = headers.get("ΕΠΩΝΥΜΟ")
        col_onoma = headers.get("ΟΝΟΜΑ")
        col_asm = headers.get("ΑΣΜ")

        if mode == "name" and (col_eponymo is None or col_onoma is None):
            raise ValueError("Δεν βρέθηκαν οι στήλες 'ΕΠΩΝΥΜΟ' και/ή 'ΟΝΟΜΑ'.")
        if mode == "asm" and col_asm is None:
            raise ValueError("Δεν βρέθηκε η στήλη 'ΑΣΜ'.")

        # Parse queries
        if mode == "name":
            q_norm = " ".join([p for p in str(query).replace(",", " ").split() if p]).upper()
            queries = [q_norm] if q_norm else []
        else:
            queries = parse_multi_values(query)

        if not queries:
            raise ValueError("Δεν δόθηκε τιμή αναζήτησης.")

        # Keep only columns we want (and exist)
        keep_cols = []
        for name in KEEP_COLUMNS:
            if name.upper() in REMOVE_COLUMNS:
                continue
            if name.upper() in headers:
                keep_cols.append(headers[name.upper()])

        if not keep_cols:
            raise ValueError("Δεν βρέθηκαν οι ζητούμενες στήλες στο αρχείο.")

        # Map out headers -> output col index
        out_header_map = {}
        for j, c in enumerate(keep_cols, start=1):
            hdr = safe_str(ws.cell(row=1, column=c).value).upper()
            if hdr:
                out_header_map[hdr] = j

        # Date columns (any header containing 'ΗΜ/ΝΙΑ')
        date_out_cols = {j for hdr, j in out_header_map.items() if "ΗΜ/ΝΙΑ" in hdr}

        # Collect matched rows
        matched_rows = []
        matched_by_key = {q.upper(): [] for q in queries} if mode == "asm" else {}

        for r in range(2, ws.max_row + 1):
            if mode == "name":
                epon = safe_str(ws.cell(row=r, column=col_eponymo).value).upper()
                onom = safe_str(ws.cell(row=r, column=col_onoma).value).upper()
                if (epon + " " + onom).strip() == queries[0]:
                    matched_rows.append(r)
            else:
                asm_val = safe_str(ws.cell(row=r, column=col_asm).value).upper()
                if asm_val in matched_by_key:
                    matched_by_key[asm_val].append(r)

        if mode == "asm":
            # Flatten in the order of queries
            for q in queries:
                matched_rows.extend(matched_by_key.get(q.upper(), []))

        if not matched_rows:
            raise ValueError("Δεν βρέθηκαν γραμμές που να ταιριάζουν.")

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Ανάλυση"

        # Writer cursor
        r_out = 1

        def write_person_block(src_rows):
            nonlocal r_out

            # Person name from first row
            first = src_rows[0]
            epon = safe_str(ws.cell(row=first, column=col_eponymo).value) if col_eponymo else ""
            onom = safe_str(ws.cell(row=first, column=col_onoma).value) if col_onoma else ""
            asm = safe_str(ws.cell(row=first, column=col_asm).value) if col_asm else ""
            person_title = (epon + " " + onom).strip()
            if asm:
                person_title = f"{person_title} ({asm})".strip()

            # 1) Name row (not merged)
            name_cell = out_ws.cell(row=r_out, column=1, value=person_title)
            name_cell.font = Font(bold=True)
            name_cell.alignment = Alignment(horizontal="left")
            r_out += 1

            # 2) Headers
            header_row = r_out
            for j, c in enumerate(keep_cols, start=1):
                out_ws.cell(row=header_row, column=j, value=ws.cell(row=1, column=c).value)
            r_out += 1

            # 3) Data rows
            data_start = r_out
            for src_r in src_rows:
                for j, c in enumerate(keep_cols, start=1):
                    val = ws.cell(row=src_r, column=c).value
                    if j in date_out_cols:
                        val = to_date(val)
                        cell = out_ws.cell(row=r_out, column=j, value=val)
                        if isinstance(val, (date, datetime)):
                            cell.number_format = "DD/MM/YYYY"
                    else:
                        out_ws.cell(row=r_out, column=j, value=val)
                r_out += 1

            data_end = r_out - 1

            # 4) Totals row
            total_row = r_out
            label_col = out_header_map.get("ΗΜ/ΝΙΑ ΕΝΑΡΞΗΣ", 1)
            label_cell = out_ws.cell(row=total_row, column=label_col, value="Σύνολο")
            label_cell.font = Font(bold=True)

            for nm in NUM_SUM_COLUMNS:
                c_idx = out_header_map.get(nm.upper())
                if c_idx:
                    start = data_start
                    end = data_end
                    if end >= start:
                        cell = out_ws.cell(
                            row=total_row,
                            column=c_idx,
                            value=f"=SUM({get_column_letter(c_idx)}{start}:{get_column_letter(c_idx)}{end})",
                        )
                    else:
                        cell = out_ws.cell(row=total_row, column=c_idx, value=0)
                    cell.font = Font(bold=True)

            # 4b) Balance row (Υπόλοιπο) κάτω ακριβώς από το Σύνολο
            balance_row = total_row + 1
            bal_label = out_ws.cell(row=balance_row, column=label_col, value="Υπόλοιπο")
            bal_label.font = Font(bold=True)

            # Υπόλοιπο = βάση - σύνολο (για ΚΑ/ΤΑΠ/ΕΛΔΥΚ)
            for nm in NUM_SUM_COLUMNS:
                c_idx = out_header_map.get(nm.upper())
                if not c_idx:
                    continue

                if nm.upper() in base_allowance:
                    total_cell_ref = f"{get_column_letter(c_idx)}{total_row}"
                    cell = out_ws.cell(
                        row=balance_row,
                        column=c_idx,
                        value=f"={base_allowance[nm.upper()]}-{total_cell_ref}",
                    )
                    cell.font = Font(bold=True)
                else:
                    out_ws.cell(row=balance_row, column=c_idx, value=None)

            # 5) Border (outline) around header+data+total+balance
            block_min_row = header_row
            block_max_row = balance_row
            block_min_col = 1
            block_max_col = len(keep_cols)
            apply_outline_border(out_ws, block_min_row, block_max_row, block_min_col, block_max_col)

            # blank line between persons
            r_out = balance_row + 2

        if mode == "asm":
            missing = []
            for q in queries:
                rows_for_key = matched_by_key.get(q.upper(), [])
                if not rows_for_key:
                    missing.append(q)
                    continue
                write_person_block(rows_for_key)

            if missing:
                messagebox.showwarning(APP_TITLE, "Δεν βρέθηκαν ΑΣΜ στο αρχείο για: " + ", ".join(missing))
        else:
            write_person_block(matched_rows)

        # Auto widths
        max_row = out_ws.max_row
        for col in range(1, len(keep_cols) + 1):
            max_len = 0
            for r in range(1, max_row + 1):
                v = out_ws.cell(row=r, column=col).value
                v = "" if v is None else str(v)
                max_len = max(max_len, len(v))
            out_ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 45)

        # Save As dialog
        fname = "report_adeies.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Αποθήκευση ως...",
            initialfile=fname,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not save_path:
            return None
        out_wb.save(save_path)
        return save_path


if __name__ == "__main__":
    app = App()
    app.mainloop()
