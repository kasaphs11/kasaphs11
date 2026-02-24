import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

# --------- Greek date formatting ----------
WEEKDAYS_GR = ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]
MONTHS_GR   = ["Ιανουαρίου","Φεβρουαρίου","Μαρτίου","Απριλίου","Μαΐου","Ιουνίου",
               "Ιουλίου","Αυγούστου","Σεπτεμβρίου","Οκτωβρίου","Νοεμβρίου","Δεκεμβρίου"]

def format_date_gr(dt: datetime) -> str:
    wd = WEEKDAYS_GR[dt.weekday()]
    m  = MONTHS_GR[dt.month-1]
    return f"{wd}, {dt.day} {m} {dt.year}"

# --------- Helpers ----------
def to_datetime_safe(x):
    if isinstance(x, datetime):
        return x
    if pd.isna(x):
        return None
    try:
        if isinstance(x, (int, float)):
            return pd.to_datetime(x, unit="D", origin="1899-12-30").to_pydatetime()
        return pd.to_datetime(str(x), dayfirst=True, errors="coerce").to_pydatetime()
    except Exception:
        return None

def parse_dates_input(s: str):
    """Δέχεται: 30/10/2025-02/11/2025 ή 30/10/2025, 01/11/2025 κ.λπ. Επιστρέφει ταξινομημένη λίστα ημερομηνιών."""
    if not s.strip():
        return []
    parts = [p.strip() for p in s.split(",") if p.strip()]
    out = []
    for p in parts:
        if "-" in p:
            a, b = [x.strip() for x in p.split("-", 1)]
            da = pd.to_datetime(a, dayfirst=True)
            db = pd.to_datetime(b, dayfirst=True)
            if pd.isna(da) or pd.isna(db):
                continue
            if da > db:
                da, db = db, da
            cur = da.normalize()
            end = db.normalize()
            while cur <= end:
                out.append(cur.to_pydatetime())
                cur += pd.Timedelta(days=1)
        else:
            d = pd.to_datetime(p, dayfirst=True)
            if not pd.isna(d):
                out.append(d.normalize().to_pydatetime())
    return sorted({d for d in out})

# --------- Layout builders ----------
def build_day_block(ws, start_row, date_header, dep_rows, ret_rows):
    """
    Γράφει ΕΝΑ ημερήσιο block (ημερομηνία, αναχωρήσεις, επιστροφές) ξεκινώντας στη γραμμή start_row.
    Επιστρέφει την επόμενη διαθέσιμη γραμμή.
    ΤΩΡΑ με στήλη Α/Α στη B (B..F = Α/Α, Επώνυμο, Όνομα, Λόχος, Τόπος).
    """
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")

    r = start_row

    # Ημερομηνία (merge B..F)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=date_header)
    c.font = Font(bold=True); c.alignment = left
    r += 2

    # --- ΑΝΑΧΩΡΗΣΕΙΣ ---
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value="ΑΝΑΧΩΡΗΣΕΙΣ ΑΔΕΙΟΥΧΩΝ ΜΕ ΠΡΟΑΝΑΧΩΡΗΣΗ")
    c.font = bold; c.alignment = center; c.fill = hdr_fill
    r += 1

    # Header row: Β=Α/Α, C..F τα πεδία
    headers = ["Α/Α", "ΕΠΩΝΥΜΟ", "ΟΝΟΜΑ", "ΛΟΧΟΣ", "ΤΟΠΟΣ ΑΔΕΙΑΣ"]
    for idx, h in enumerate(headers, start=2):  # 2..6
        ws.cell(row=r, column=idx, value=h).font = bold
        ws.cell(row=r, column=idx).alignment = center
        ws.cell(row=r, column=idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=idx).fill = PatternFill("solid", fgColor="F2F2F2")
    r += 1

    # Data rows με αρίθμηση που ξεκινά από 1
    for i, vals in enumerate(dep_rows, start=1):
        # Α/Α
        ws.cell(row=r, column=2, value=i).alignment = center
        ws.cell(row=r, column=2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        # Επώνυμο, Όνομα, Λόχος, Τόπος -> C..F
        for j, v in enumerate(vals, start=3):
            ws.cell(row=r, column=j, value=v)
            ws.cell(row=r, column=j).alignment = left if j in (3,4) else center
            ws.cell(row=r, column=j).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        r += 1

    r += 1  # κενό

    # --- ΕΠΙΣΤΡΟΦΕΣ ---
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value="ΕΠΙΣΤΡΟΦΕΣ ΑΔΕΙΟΥΧΩΝ")
    c.font = bold; c.alignment = center; c.fill = hdr_fill
    r += 1

    for idx, h in enumerate(headers, start=2):
        ws.cell(row=r, column=idx, value=h).font = bold
        ws.cell(row=r, column=idx).alignment = center
        ws.cell(row=r, column=idx).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=idx).fill = PatternFill("solid", fgColor="F2F2F2")
    r += 1

    # Data rows για επιστροφές
    for i, vals in enumerate(ret_rows, start=1):
        ws.cell(row=r, column=2, value=i).alignment = center
        ws.cell(row=r, column=2).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for j, v in enumerate(vals, start=3):
            ws.cell(row=r, column=j, value=v)
            ws.cell(row=r, column=j).alignment = left if j in (3,4) else center
            ws.cell(row=r, column=j).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        r += 1

    r += 2  # 2 κενές γραμμές πριν το επόμενο block
    return r

def build_notes(ws, start_row, rows=8):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = start_row
    # Τίτλος "Παρατηρήσεις" (merge B..F)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value="ΠΑΡΑΤΗΡΗΣΕΙΣ:")
    c.font = bold; c.alignment = center
    r += 1

# --------- Core run ----------
def run():
    try:
        src = var_source.get().strip()
        dates_s = var_dates.get().strip()
        start_col = var_start_col.get().strip()
        end_col   = var_end_col.get().strip()
        cols_text = var_cols.get().strip()

        if not os.path.exists(src):
            messagebox.showerror("Σφάλμα", "Δώσε σωστό πηγαίο Excel.")
            return

        days = parse_dates_input(dates_s)
        if not days:
            messagebox.showerror("Ημερομηνίες", "Δώσε τουλάχιστον μία έγκυρη ημερομηνία.")
            return

        df = pd.read_excel(src, sheet_name=0)
        if start_col not in df.columns or end_col not in df.columns:
            messagebox.showerror(
                "Στήλες",
                f"Δεν βρέθηκαν οι στήλες: {start_col} / {end_col}\nΔιαθέσιμες:\n{list(df.columns)}"
            )
            return

        df["_start"] = pd.to_datetime(df[start_col].apply(to_datetime_safe)).dt.normalize()
        df["_end"]   = pd.to_datetime(df[end_col].apply(to_datetime_safe)).dt.normalize()

        # Τα πεδία από το source (η σειρά τους: ΕΠΩΝΥΜΟ, ΟΝΟΜΑ, ΛΟΧΟΣ, ΤΟΠΟΣ ΜΕΤΑΒΑΣΗΣ)
        col_names = [c.strip() for c in cols_text.split(",") if c.strip()]

        wb = Workbook()
        ws = wb.active
        ws.title = "Άδειες"

        # --------- Page setup (Print settings) ----------
        ws.page_setup.scale = 90
        


        # Πλάτη στηλών: B..F (Α/Α, ΕΠΩΝΥΜΟ, ΟΝΟΜΑ, ΛΟΧΟΣ, ΤΟΠΟΣ)
        widths = [5, 22, 28, 13, 24]
        for i, w in enumerate(widths, start=2):  # 2..6
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.column_dimensions["A"].width = 2

        r = 1
        total_dep = 0
        total_ret = 0

        for X in days:
            dep_date = (X + timedelta(days=1)).date()
            dep_df = df[df["_start"] == pd.Timestamp(dep_date)]
            ret_df = df[df["_end"]   == pd.Timestamp(X.date())]

            dep_rows = [[row.get(c, "") for c in col_names] for _, row in dep_df.iterrows()]
            ret_rows = [[row.get(c, "") for c in col_names] for _, row in ret_df.iterrows()]

            header = format_date_gr(X)
            r = build_day_block(ws, r, header, dep_rows, ret_rows)

            total_dep += len(dep_rows)
            total_ret += len(ret_rows)

        # Παρατηρήσεις στο τέλος
        build_notes(ws, r, rows=8)

        # Προτείνεται όνομα αρχείου
        first_date = days[0]
        last_date  = days[-1]
        if len(days) == 1:
            default_name = f"ΕΠΟΠΤΕΙΟ {first_date.strftime('%d-%m-%Y')}.xlsx"
        else:
            default_name = f"ΕΠΟΠΤΕΙΟ {first_date.strftime('%d')},{last_date.strftime('%d-%m-%Y')}.xlsx"

        save_path = filedialog.asksaveasfilename(
            title="Αποθήκευση ως",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name
        )
        if not save_path:
            return

        wb.save(save_path)

        messagebox.showinfo(
            "OK",
            f"Δημιουργήθηκε αρχείο με {len(days)} ημέρες.\n"
            f"Σύνολο Αναχωρήσεων: {total_dep}\n"
            f"Σύνολο Επιστροφών: {total_ret}"
        )

    except PermissionError:
        messagebox.showerror("Αποθήκευση", "Δεν μπόρεσα να αποθηκεύσω (είναι ανοιχτό το αρχείο;)")
    except Exception as e:
        messagebox.showerror("Σφάλμα", str(e))

# --------- GUI ----------
root = tk.Tk()
root.title("Άδειες: Πολλαπλές Ημερομηνίες (χωρίς template)")

p = {"padx": 6, "pady": 6}
var_source = tk.StringVar()
var_dates  = tk.StringVar()
var_start_col = tk.StringVar(value="ΗΜ/ΝΙΑ ΕΝΑΡΞΗΣ")
var_end_col   = tk.StringVar(value="ΗΜ/ΝΙΑ ΛΗΞΗΣ")
var_cols      = tk.StringVar(value="ΕΠΩΝΥΜΟ,ΟΝΟΜΑ,ΛΟΧΟΣ,ΤΟΠΟΣ ΜΕΤΑΒΑΣΗΣ")

frm = ttk.Frame(root); frm.grid(row=0, column=0, sticky="nsew")
root.columnconfigure(0, weight=1); root.rowconfigure(0, weight=1)

def pick_src():
    p = filedialog.askopenfilename(title="Πηγαίο Excel", filetypes=[("Excel", "*.xlsx *.xls")])
    if p: var_source.set(p)

r=0
ttk.Label(frm, text="Πηγαίο Excel:").grid(row=r, column=0, sticky="w", **p)
ttk.Entry(frm, textvariable=var_source, width=66).grid(row=r, column=1, **p)
ttk.Button(frm, text="Επιλογή...", command=pick_src).grid(row=r, column=2, **p); r+=1

ttk.Label(frm, text="Ημερομηνίες (π.χ. 30/10/2025-02/11/2025 ή 30/10/2025, 1/11/2025):").grid(row=r, column=0, sticky="w", **p)
ttk.Entry(frm, textvariable=var_dates, width=66).grid(row=r, column=1, **p); r+=1

ttk.Label(frm, text="Στήλη Ημερομηνίας Έναρξης:").grid(row=r, column=0, sticky="w", **p)
ttk.Entry(frm, textvariable=var_start_col).grid(row=r, column=1, sticky="we", **p); r+=1

ttk.Label(frm, text="Στήλη Ημερομηνίας Λήξης:").grid(row=r, column=0, sticky="w", **p)
ttk.Entry(frm, textvariable=var_end_col).grid(row=r, column=1, sticky="we", **p); r+=1

ttk.Label(frm, text="Στήλες προς αντιγραφή (comma):").grid(row=r, column=0, sticky="w", **p)
ttk.Entry(frm, textvariable=var_cols).grid(row=r, column=1, sticky="we", **p); r+=1

ttk.Button(frm, text="Εκτέλεση", command=run).grid(row=r, column=1, pady=10)

root.mainloop()
