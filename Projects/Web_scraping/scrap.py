import os
import re
import time
import urllib.parse
from typing import Optional, Tuple

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# =========================
# CONFIG
# =========================
IN_PATH  = "ΤΙΜΕΣ.xlsx"
OUT_PATH = "ΤΙΜΕΣ_ΕΛΛΑΔΑ_bestprice.xlsx"   # ενημερώνεται συνέχεια (in-place)

HEADLESS        = True
SLOW_MO_MS      = 0
PAGE_TIMEOUT_MS = 60_000

SLEEP_BETWEEN_PRODUCTS = 0.6
AUTOSAVE_EVERY         = 50

MAX_TO_PROCESS: Optional[int] = None   # None = όλα, π.χ. 50 για test
MAX_RETRIES = 2                        # retries για TIMEOUT/ERROR

SIMILARITY_THRESHOLD = 0.30           # κάτω από αυτό → LOW_SIM

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)

BESTPRICE_POTA_CATEGORY_URL = "https://www.bestprice.gr/cat/2385/pota.html"
PRICE_RE = re.compile(r"(\d+(?:[.,]\d{1,2})?)\s*€")

PRODUCT_CARD_SELECTORS = [
    "div[class*='product-card']",
    "article[class*='product']",
    "li[class*='product']",
    "div[class*='ProductCard']",
    "div[class*='product_card']",
    "div[data-testid*='product']",
]

WAIT_SELECTOR         = "div[class*='product'], article[class*='product'], li[class*='product']"
WAIT_SELECTOR_TIMEOUT = 8_000


# =========================
# NORMALIZE
# =========================
def normalize_query_only_name(name: str) -> str:
    s = (name or "").strip()
    s = s.replace("％", "%").replace("﹪", "%").replace("٪", "%")
    s = re.sub(r"\b\d+[.,]\d+\s*(?:l|lt)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b\d+(?:[.,]\d+)?\s*%\s*(?:vol)?\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b(?:alc|alcohol|abv)\s*\d+(?:[.,]\d+)?\s*%\b", " ", s, flags=re.IGNORECASE)
    s = s.replace(",", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\b\d+(?:[.,]\d+)?\s*(?:cl|ml|l|lt)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b(y\.?o\.?|yo|years?\s*old)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\b(vol|proof)\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[^A-Za-z0-9Α-Ωα-ω\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def short_query(full_query: str, words: int = 2) -> str:
    tokens = full_query.split()
    return " ".join(tokens[:words]) if len(tokens) >= words else full_query


# =========================
# SIMILARITY
# =========================
def token_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    ta = set(re.findall(r"[A-Za-z0-9Α-Ωα-ω]+", a.lower()))
    tb = set(re.findall(r"[A-Za-z0-9Α-Ωα-ω]+", b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


# =========================
# VOLUME EXTRACTION
# =========================
_VOL_RE = re.compile(r"\b(\d+(?:[.,]\d+)?)\s*(ml|cl|lt?)\b", flags=re.IGNORECASE)

def extract_volume_ml(text: str) -> Optional[int]:
    """
    Εξάγει τον πρώτο όγκο από τίτλο και τον επιστρέφει σε ml.
    700ml→700, 70cl→700, 1L→1000, 0,7l→700, 1lt→1000
    Επιστρέφει None αν δεν βρεθεί.
    """
    m = _VOL_RE.search(text or "")
    if not m:
        return None
    value = float(m.group(1).replace(",", "."))
    unit  = m.group(2).lower()
    if unit == "cl":
        return int(round(value * 10))
    if unit in ("l", "lt", "ltr"):
        return int(round(value * 1000))
    return int(round(value))  # ml


# =========================
# SCRAPER — CORE
# =========================
def parse_price_from_text(text: str) -> Optional[float]:
    m = PRICE_RE.search(text)
    if not m:
        return None
    return float(m.group(1).replace(",", "."))


def absolutize_bestprice_url(base: str, href: str) -> str:
    if not href:
        return base
    if href.startswith("http"):
        return href
    if href.startswith("/"):
        return "https://www.bestprice.gr" + href
    return base.rstrip("/") + "/" + href.lstrip("/")


def _wait_for_products(page) -> None:
    try:
        page.wait_for_selector(WAIT_SELECTOR, timeout=WAIT_SELECTOR_TIMEOUT)
    except PlaywrightTimeoutError:
        pass


def _scrape_first_product(page, search_url: str) -> Tuple[Optional[float], str, str]:
    """Επιστρέφει (price, link, product_title)."""
    js = """
    (cardSelectors) => {
        const priceRe = /(\\d+(?:[.,]\\d{1,2})?)\\s*€/;

        function extractFromElement(el) {
            const text = (el.innerText || "").replace(/\\s+/g, " ").trim();
            const m = text.match(priceRe);
            if (!m) return null;
            const a = el.querySelector('a[href*=".html"], a[href*="/p/"]');
            const href = a ? (a.getAttribute("href") || "") : "";
            const titleEl = el.querySelector('h2, h3, [class*="title"], [class*="name"], [class*="Title"]');
            const title = titleEl
                ? (titleEl.innerText || "").trim()
                : (a ? (a.innerText || "").trim() : "");
            return { priceText: m[0], href, title };
        }

        // 1) Targeted selectors
        for (const sel of cardSelectors) {
            const cards = Array.from(document.querySelectorAll(sel));
            for (const card of cards) {
                const r = extractFromElement(card);
                if (r) return r;
            }
        }

        // 2) Fallback generic scan
        const root = document.querySelector("main") || document.body;
        const blocks = Array.from(root.querySelectorAll("article, li, div, section"));
        for (const b of blocks) {
            if (b.children.length > 20) continue;
            const r = extractFromElement(b);
            if (r) return r;
        }
        return null;
    }
    """

    result = page.evaluate(js, PRODUCT_CARD_SELECTORS)

    if not result:
        return None, "", ""

    price = parse_price_from_text(result["priceText"])
    href  = result.get("href") or ""
    title = result.get("title") or ""

    if price is None:
        return None, "", title

    link = absolutize_bestprice_url(search_url, href) if href else search_url
    return price, link, title


def bestprice_search(page, query: str) -> Tuple[Optional[float], str, str, float]:
    """Επιστρέφει (price, link, matched_title, similarity_score)."""
    url = BESTPRICE_POTA_CATEGORY_URL + "?" + urllib.parse.urlencode({"q": query})
    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT_MS)
    _wait_for_products(page)
    price, link, title = _scrape_first_product(page, url)
    sim = token_similarity(query, title) if title else 0.0
    return price, link, title, sim


# =========================
# SMART SEARCH
# =========================
def smart_search(page, raw_name: str) -> Tuple[Optional[float], str, Optional[int], str]:
    """
    Επιστρέφει (price, link, volume_ml, notes).
    Στρατηγική: full query → short query (2 λέξεις) αν NONE ή low sim.
    """
    full_q = normalize_query_only_name(raw_name)

    # Pass 1: full query
    price, link, title, sim = bestprice_search(page, full_q)
    vol = extract_volume_ml(title)

    if price is not None and sim >= SIMILARITY_THRESHOLD:
        notes = f"OK | {vol or '?'}ml | sim={sim:.2f} | title='{title}' | q='{full_q}'"
        return price, link, vol, notes

    # Pass 2: short query
    short_q = short_query(full_q, words=2)
    if short_q and short_q != full_q:
        price2, link2, title2, sim2 = bestprice_search(page, short_q)
        vol2 = extract_volume_ml(title2)

        if price2 is not None and sim2 >= SIMILARITY_THRESHOLD:
            notes = f"OK(short) | {vol2 or '?'}ml | sim={sim2:.2f} | title='{title2}' | q='{short_q}'"
            return price2, link2, vol2, notes

        if price2 is not None:
            notes = f"LOW_SIM(short) | {vol2 or '?'}ml | sim={sim2:.2f} | title='{title2}' | q='{short_q}'"
            return price2, link2, vol2, notes

    # Full query είχε τιμή αλλά low sim
    if price is not None:
        notes = f"LOW_SIM | {vol or '?'}ml | sim={sim:.2f} | title='{title}' | q='{full_q}'"
        return price, link, vol, notes

    # Τίποτα
    notes = f"NONE | q='{full_q}'"
    return None, "", None, notes


# =========================
# EXCEL UTILS
# =========================
def ensure_col(ws, headers: list, title: str) -> int:
    for i, h in enumerate(headers):
        if h == title:
            return i + 1
    headers.append(title)
    ws.cell(row=1, column=len(headers), value=title)
    return len(headers)


# =========================
# MAIN
# =========================
def main():
    source_path = OUT_PATH if os.path.exists(OUT_PATH) else IN_PATH
    target_path = OUT_PATH

    wb = load_workbook(source_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    c_name    = ensure_col(ws, headers, "Ονομασία Εμπορεύματος")
    c_gr      = ensure_col(ws, headers, "ΕΛΛΑΔΑ")
    c_gr_link = ensure_col(ws, headers, "GR_link")
    c_gr_ml   = ensure_col(ws, headers, "GR_ml")       # ← νέα στήλη
    c_notes   = ensure_col(ws, headers, "Σημειώσεις")

    processed = 0
    skipped   = 0
    updated   = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=SLOW_MO_MS)
        context = browser.new_context(user_agent=USER_AGENT, locale="el-GR")
        page    = context.new_page()

        for r in range(2, ws.max_row + 1):

            # Skip αν έχει ήδη τιμή
            current_price = ws.cell(r, c_gr).value
            if current_price is not None and str(current_price).strip() not in ("", "0"):
                try:
                    if float(str(current_price).replace(",", ".")) > 0:
                        skipped += 1
                        continue
                except ValueError:
                    pass

            name = ws.cell(r, c_name).value
            if not name:
                continue

            raw = str(name).strip()

            # Retry loop
            price = link = vol = notes = None
            for attempt in range(1, MAX_RETRIES + 2):
                try:
                    price, link, vol, notes = smart_search(page, raw)
                    break
                except PlaywrightTimeoutError:
                    notes = f"TIMEOUT(attempt {attempt})"
                    print(f"  [TIMEOUT] attempt {attempt}/{MAX_RETRIES + 1} | {raw}")
                    if attempt <= MAX_RETRIES:
                        time.sleep(2 * attempt)
                except Exception as e:
                    notes = f"ERROR {type(e).__name__}(attempt {attempt})"
                    print(f"  [ERROR] attempt {attempt}/{MAX_RETRIES + 1} | {raw} -> {type(e).__name__}: {e}")
                    if attempt <= MAX_RETRIES:
                        time.sleep(2 * attempt)

            # Γράψε στο Excel
            ws.cell(r, c_gr).value      = price
            ws.cell(r, c_gr_link).value = link if link else None
            ws.cell(r, c_gr_ml).value   = vol          # ml ή None αν δεν βρέθηκε
            ws.cell(r, c_notes).value   = notes or "ERROR_UNKNOWN"

            if price is not None:
                updated += 1
                status = "OK" if "LOW_SIM" not in (notes or "") else "LOW_SIM"
                vol_str = f"{vol}ml" if vol else "?ml"
                print(f"[{status}] {raw} -> {price:.2f}€ ({vol_str}) | {notes}")
            else:
                print(f"[NONE] {raw} | {notes}")

            processed += 1
            time.sleep(SLEEP_BETWEEN_PRODUCTS)

            if processed % AUTOSAVE_EVERY == 0:
                wb.save(target_path)
                print(f"--- autosaved ({processed} processed, {updated} updated) ---")

            if MAX_TO_PROCESS is not None and processed >= MAX_TO_PROCESS:
                break

        browser.close()

    wb.save(target_path)
    print(f"\nDone. updated={updated}, skipped(already filled)={skipped}, processed(now)={processed}")
    print("Saved:", target_path)


if __name__ == "__main__":
    main()
